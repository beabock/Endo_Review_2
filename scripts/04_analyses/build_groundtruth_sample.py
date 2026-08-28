#!/usr/bin/env python3
# BMB 2026-08-28
# Builds the manually-curated ground-truth / validation sample for the extraction
# pipeline (NPH-MS-2026-57711 resubmission, Referee 2). See NPH_task2_ground_truth_plan.md
# and NPH_extraction_schema_design.md.
#
#   1. Builds a paper-level frame from the extraction output + abstract metadata,
#      assigning stratification variables (doc type, function, continent, era).
#   2. Draws a stratified random sample (fixed seed): a core sample + a targeted
#      oversample of rare functional categories, in a stratified ORDER so that any
#      prefix is itself balanced (the kappa block can be extended later).
#   3. Splits into: a kappa block (all annotators) + solo blocks (Bea + a 4th reviewer).
#      Nancy & Kitty rate only the kappa block.
#   4. Writes one .xlsx workbook per person. One row per (paper, reading stage): a
#      full-text paper gets an abstract-only row THEN a full-text row (adjacent, same
#      reader) so we also get a human measure of what abstracts omit vs their own full
#      text. Dropdown-validated, 5 interaction slots, method tick-columns.
#
# BLIND annotation: no model output is shown (avoids anchoring). score_groundtruth.py
# merges the completed sheets with an extraction run for precision / recall / F1 / kappa.
#
# Usage:
#   python scripts/04_analyses/build_groundtruth_sample.py
#   python scripts/04_analyses/build_groundtruth_sample.py --n-kappa 40 --r4-solo 70

from __future__ import annotations

import argparse
import sys
from collections import Counter, defaultdict
from importlib import util
from pathlib import Path

import numpy as np
import pandas as pd

ROOT = Path(__file__).resolve().parents[2]
UTILS = ROOT / "scripts" / "utils"

_spec = util.spec_from_file_location("country_mapping", UTILS / "country_mapping.py")
country_mapping = util.module_from_spec(_spec)
_spec.loader.exec_module(country_mapping)
get_continent = country_mapping.get_continent

DEFAULT_EXTRACTION = ROOT / "data" / "Ollama_cleaned_synresolved_standardized_year.csv"
ABSTRACT_META = ROOT / "data" / "Abstracts" / "Abstracts_for_Monsoon.csv"
ABSTRACT_DEDUPED = ROOT / "data" / "Abstracts" / "All_abstracts_deduped.csv"
OUT_DIR = ROOT / "results" / "manual_validation" / "groundtruth"

# annotator key -> display label.
ANNOTATORS = {
    "bea": "Bea",
    "nancy": "Nancy",
    "kitty": "Kitty",
    "ian": "Ian",
    "jack": "Jack",
}
# reviewers who take a solo block in addition to the kappa block (besides Bea, who
# takes the remainder). Nancy & Kitty do the kappa block only.
EXTRA_REVIEWERS = ["ian", "jack"]

# --- controlled vocabularies -------------------------------------------------------
# Kept in sync with guild_rubric.md. Written to a hidden "Lists" sheet and range-
# referenced by the dropdowns (so values may contain commas, and the lists stay
# editable). Validation is set to "warning" style, so an annotator can also type a
# value that isn't in the list (e.g. a specific language / biome).
VOCAB = {
    "yn": ["yes", "no"],
    "yn_unclear": ["yes", "no", "unclear"],
    "paper_reviewed": [
        "complete",
        "could not access the paper (full text needed)",
        "review or secondary compilation - no primary data of its own",
        "not a fungus-in-plant study",
    ],
    "primary_aim": [
        "endophyte ecology / diversity / community",
        "endophyte function: growth, nutrition, or stress tolerance",
        "endophyte function: disease protection / biocontrol",
        "natural products / secondary metabolites / bioprospecting",
        "plant pathology (the fungus studied as a pathogen)",
        "fungal taxonomy / systematics / new species",
        "genomics / transcriptomics / methods development",
        "other (type it)",
    ],
    "sampling_approach": [
        "one or a few isolates (named)", "culture survey (many isolates, named)",
        "community metabarcoding / NGS (OTU or ASV table)",
        "direct observation only (microscopy / histology, no isolation or sequencing)",
        "mixed methods (type which)",
        "review / synthesis / secondary compilation - no new fungi obtained here",
        "other (type it)", "unclear",
    ],
    "language": [
        "English",
        "non-English (type which)",
        "English + non-English (bilingual abstract, or English abstract + non-English body)",
        "unclear",
    ],
    "sampling_location_status": [
        "sampling location is explicitly stated",
        "sampling location is inferable from a named field site / region",
        "no sampling location - only the authors' institutional country is given",
        "no geographic information at all",
    ],
    "strain_variation": [
        "yes - noted", "no - not noted", "only one strain / isolate studied",
    ],
    "biome": [
        "tropical / subtropical moist forest", "tropical / subtropical dry forest",
        "temperate or boreal forest", "Mediterranean forest / woodland / scrub",
        "tropical grassland / savanna / shrubland",
        "temperate grassland / steppe / prairie", "montane grassland / alpine",
        "desert / xeric shrubland", "wetland / flooded grassland / bog",
        "mangrove", "tundra / polar", "marine / coastal",
        "agriculture / cultivated (crop, orchard, plantation, sown pasture)",
        "urban / garden / botanical collection", "greenhouse / growth chamber",
        "other (type it)", "not stated",
    ],
    "tissue": [
        "leaf", "root", "stem / wood / bark", "seed", "fruit",
        "flower / reproductive", "whole plant", "other (type it)", "not stated",
    ],
    # fungal lifestyle / trophic mode - how the fungus makes its living in this
    # interaction. Values follow FungalTraits primary_lifestyle (Polme et al. 2020);
    # the three endophyte types are grouped (the tissue column records leaf vs root).
    "fungal_lifestyle": [
        "endophyte (foliar, root, or dark-septate; symptomless in living tissue)",
        "latent pathogen / hemibiotroph",
        "plant pathogen",
        "wood saprotroph",
        "litter saprotroph",
        "soil saprotroph",
        "unspecified / other saprotroph",
        "arbuscular mycorrhizal",
        "ectomycorrhizal",
        "ericoid mycorrhizal",
        "orchid mycorrhizal",
        "mycoparasite / antagonist of other microbes",
        "epiphyte (surface only)",
        "lichenised",
        "context-dependent (varies by host / environment / strain)",
        "not clear from the paper",
    ],
    # net outcome for THIS host plant, as the paper reports it
    "effect_on_host": [
        "not reported (fungus just isolated / detected)",
        "no visible symptoms; benefit or harm not tested",
        "neutral / commensal (plant response tested; no net effect)",
        "beneficial to the plant",
        "harmful to the plant (disease / reduced fitness)",
        "context-dependent (varies by host / environment / strain)",
        "unclear",
    ],
    "evidence_basis": [
        "experimentally tested in this study",
        "observed / measured in this study (no manipulation)",
        "inferred from what the fungus usually does (taxon reputation / database)",
        "asserted with no support",
        "not stated",
    ],
    "sterilisation_checked": [
        "yes - a control was done (final rinse plated, tissue imprints, or similar)",
        "no - standard sterilisation, no check described",
        "not a culture-based study",
        "cannot tell from the text",
    ],
}

N_SLOTS = 5
# Method columns: plain-English headers (no "m_" prefix - Kitty / Nancy found the codes
# confusing). Tick "yes" in every one that applies. Full definitions in guild_rubric.md.
METHOD_COLS = [
    "culture from sterilised tissue",
    "microscopy in tissue",
    "direct sequencing from tissue",
    "isolate ID by sequencing",
    "resynthesis / re-inoculation",
    "method not stated",
]

GUILD_TO_FUNCTION = {
    "endophyte": "endophyte_asymptomatic", "endophytic": "endophyte_asymptomatic",
    "pathogen": "pathogen", "pathogenic": "pathogen", "phytopathogen": "pathogen",
    "mycorrhiza": "mycorrhizal", "mycorrhizal": "mycorrhizal", "ectomycorrhiza": "mycorrhizal",
    "mutualist": "mutualist_pgpr", "symbiotic": "mutualist_pgpr", "symbiont": "mutualist_pgpr",
    "pgpr": "mutualist_pgpr",
    "saprotroph": "saprotroph", "saprobic": "saprotroph", "decomposer": "saprotroph",
    "biocontrol": "biocontrol_antagonist", "antagonist": "biocontrol_antagonist",
    "antifungal": "biocontrol_antagonist",
}
RARE_FUNCTIONS = {"saprotroph", "mutualist_pgpr", "biocontrol_antagonist"}
RARE_KEYWORDS = [
    "commensal", "latent pathogen", "symptomless", "asymptomatic then",
    "quiescent", "endophyte to pathogen", "life-cycle transition",
]


def norm_doi(x):
    if pd.isna(x):
        return None
    return str(x).strip().lower().lstrip("doi:").strip().replace("https://doi.org/", "")


def era_of(year):
    if pd.isna(year):
        return "unknown"
    y = int(year)
    return "pre-2000" if y < 2000 else "2000-2012" if y <= 2012 else "2013-2024"


def build_paper_frame(extraction_path: Path) -> pd.DataFrame:
    df = pd.read_csv(extraction_path, low_memory=False)
    df["doi_n"] = df["doi"].map(norm_doi)
    df["function"] = (df["primary_guild"].astype(str).str.strip().str.lower()
                      .map(GUILD_TO_FUNCTION).fillna("unknown"))
    df["continent"] = df["country"].map(
        lambda c: get_continent(c) if isinstance(c, str) and len(c) == 3 else None)

    def agg_paper(g):
        funcs = [f for f in g["function"] if f != "unknown"]
        conts = [c for c in g["continent"].dropna() if c and c != "Other"]
        ds = g["data_source"].astype(str).str.lower()
        year = g["publication_year"].dropna()
        return pd.Series({
            "doi_n": g["doi_n"].dropna().iloc[0] if g["doi_n"].notna().any() else None,
            "doc_type": "full-text" if (ds == "full-text").any() else "abstract-only",
            "function": Counter(funcs).most_common(1)[0][0] if funcs else "unknown",
            "continent": Counter(conts).most_common(1)[0][0] if conts else "unassigned",
            "year": int(year.iloc[0]) if len(year) else np.nan,
            "n_model_rows": len(g),
            "notes_blob": " ".join(
                str(x) for x in g.get("interaction_notes", pd.Series(dtype=str)).dropna()
            ).lower(),
        })

    papers = df.groupby("paper_id", dropna=True).apply(agg_paper).reset_index()
    papers["era"] = papers["year"].map(era_of)
    return papers


def attach_metadata(papers: pd.DataFrame) -> pd.DataFrame:
    meta = pd.read_csv(ABSTRACT_META, low_memory=False)
    meta["doi_n"] = meta["DOI"].map(norm_doi)
    meta = (meta.dropna(subset=["doi_n"]).drop_duplicates("doi_n")
            .rename(columns={"Title": "title", "Authors": "authors",
                             "Source.title": "journal", "Abstract": "abstract"})
            [["doi_n", "title", "authors", "journal", "abstract"]])

    lang = pd.DataFrame(columns=["doi_n", "language_hint"])
    try:
        dd = pd.read_csv(ABSTRACT_DEDUPED, low_memory=False,
                         usecols=["DOI", "Language.of.Original.Document"])
        dd["doi_n"] = dd["DOI"].map(norm_doi)
        lang = (dd.dropna(subset=["doi_n"]).drop_duplicates("doi_n")
                .rename(columns={"Language.of.Original.Document": "language_hint"})
                [["doi_n", "language_hint"]])
    except Exception as exc:  # noqa: BLE001
        print(f"  (language hint unavailable: {exc})")

    out = papers.merge(meta, on="doi_n", how="left").merge(lang, on="doi_n", how="left")
    out["doi_link"] = out["doi_n"].map(
        lambda d: f"https://doi.org/{d}" if isinstance(d, str) else "")
    return out


def proportional_allocation(counts: dict, total: int, floor: int = 2) -> dict:
    keys = [k for k, v in counts.items() if v > 0]
    base = {k: min(floor, counts[k]) for k in keys}
    remaining = total - sum(base.values())
    if remaining > 0:
        wt = sum(counts[k] for k in keys)
        for k in keys:
            base[k] += int(round(remaining * counts[k] / wt))
    diff = total - sum(base.values())
    order = sorted(keys, key=lambda k: counts[k], reverse=diff > 0)
    i = 0
    while diff != 0 and order and i < 10000:
        k = order[i % len(order)]
        step = 1 if diff > 0 else -1
        if 1 <= base[k] + step <= counts[k]:
            base[k] += step
            diff -= step
        i += 1
    return base


def stratified_sample(pool: pd.DataFrame, n: int, rng, strata_cols) -> pd.DataFrame:
    counts = {k: v for k, v in pool.groupby(strata_cols).size().to_dict().items() if v > 0}
    alloc = proportional_allocation(counts, n)
    picks = []
    for key, k_n in alloc.items():
        key_t = key if isinstance(key, tuple) else (key,)
        cell = pool[np.logical_and.reduce([pool[c] == v for c, v in zip(strata_cols, key_t)])]
        take = min(k_n, len(cell))
        if take:
            picks.append(cell.sample(n=take, random_state=rng.integers(1 << 31)))
    return pd.concat(picks).drop_duplicates("paper_id")


def stratified_order(df: pd.DataFrame, rng, strata_cols) -> pd.DataFrame:
    """Round-robin interleave strata so any prefix of the result is balanced."""
    buckets = defaultdict(list)
    for _, r in df.sample(frac=1, random_state=rng.integers(1 << 31)).iterrows():
        buckets[tuple(r[c] for c in strata_cols)].append(r)
    keys = sorted(buckets, key=lambda k: -len(buckets[k]))
    ordered, exhausted = [], False
    while not exhausted:
        exhausted = True
        for k in keys:
            if buckets[k]:
                ordered.append(buckets[k].pop(0))
                exhausted = False
    return pd.DataFrame(ordered).reset_index(drop=True)


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--extraction", type=Path, default=DEFAULT_EXTRACTION)
    ap.add_argument("--seed", type=int, default=20260828)
    ap.add_argument("--n-core", type=int, default=200)
    ap.add_argument("--n-rare", type=int, default=30)
    ap.add_argument("--full-text-frac", type=float, default=0.35)
    ap.add_argument("--n-kappa", type=int, default=50)
    ap.add_argument("--n-calibration", type=int, default=15,
                    help="first N kappa-block papers = a calibration round: everyone does "
                         "these first, then a reconciliation call refines the rubric before "
                         "the rest. Reported kappa uses the remaining measurement papers.")
    ap.add_argument("--solo-per-extra", type=int, default=50,
                    help="solo-block size for each of EXTRA_REVIEWERS (Ian, Jack); "
                         "they also rate the whole kappa block. Bea takes the remainder.")
    args = ap.parse_args()

    if not args.extraction.exists():
        print(f"ERROR: extraction file not found: {args.extraction}")
        return 1
    rng = np.random.default_rng(args.seed)
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    print(f"Building paper frame from {args.extraction.name} ...")
    papers = build_paper_frame(args.extraction)
    papers = attach_metadata(papers)
    papers["search_text"] = (papers["notes_blob"].fillna("") + " "
                             + papers["abstract"].fillna("").str.lower())
    frame_cols = ["paper_id", "doi_n", "doc_type", "function", "continent", "era",
                  "year", "n_model_rows", "journal", "language_hint"]
    papers[frame_cols].to_csv(OUT_DIR / "sample_frame.csv", index=False)
    print(f"  {len(papers):,} papers ({(papers.doc_type == 'full-text').sum():,} full-text)")

    n_ft = int(round(args.n_core * args.full_text_frac))
    ft = papers[papers.doc_type == "full-text"]
    ab = papers[papers.doc_type == "abstract-only"]
    core = pd.concat([
        stratified_sample(ft, min(n_ft, len(ft)), rng, ["function", "era"]),
        stratified_sample(ab, args.n_core - n_ft, rng, ["function", "era"]),
    ]).drop_duplicates("paper_id")
    core["bucket"] = "core"

    used = set(core.paper_id)
    rare_pool = papers[
        (papers.function.isin(RARE_FUNCTIONS)
         | papers.search_text.str.contains("|".join(RARE_KEYWORDS), regex=True))
        & ~papers.paper_id.isin(used)]
    rare = rare_pool.sample(n=min(args.n_rare, len(rare_pool)),
                            random_state=rng.integers(1 << 31))
    rare["bucket"] = "rare_function_oversample"

    # kappa block: a stratified pool of the core that keeps the core's doc-type mix
    # (so Nancy/Kitty's full-text load matches the sample, not an over-weighted share),
    # ordered so any prefix stays balanced -> raising --n-kappa later is valid.
    pool_n = min(max(2 * args.n_kappa, args.n_kappa + 20), len(core))
    core_ft_frac = (core.doc_type == "full-text").mean()
    pool_ft_n = round(pool_n * core_ft_frac)
    kappa_pool = pd.concat([
        stratified_sample(core[core.doc_type == "full-text"], pool_ft_n, rng, ["function"]),
        stratified_sample(core[core.doc_type == "abstract-only"], pool_n - pool_ft_n,
                          rng, ["function"]),
    ]).drop_duplicates("paper_id")
    kappa_pool = stratified_order(kappa_pool, rng, ["function"])
    kappa_pool["order"] = range(len(kappa_pool))
    kappa_ids = set(kappa_pool.head(args.n_kappa).paper_id)
    calibration_ids = set(kappa_pool.head(args.n_calibration).paper_id)

    sample = pd.concat([core.merge(kappa_pool[["paper_id", "order"]], on="paper_id", how="left"),
                        rare]).drop_duplicates("paper_id").reset_index(drop=True)
    sample["order"] = sample["order"].fillna(9999).astype(int)
    sample["assignment"] = np.where(sample.paper_id.isin(kappa_ids), "kappa_block", "solo")
    sample["phase"] = np.select(
        [sample.paper_id.isin(calibration_ids), sample.paper_id.isin(kappa_ids)],
        ["calibration", "measurement"], default="")

    # solo split: each EXTRA_REVIEWER takes a stratified slice of the solo block
    # (keeping the solo block's ~31% full-text mix); Bea takes the remainder.
    sample["annotator"] = np.where(sample.assignment == "kappa_block", "ALL", "bea")
    remaining = set(sample[sample.annotator == "bea"].paper_id)
    for who in EXTRA_REVIEWERS:
        pool = sample[sample.paper_id.isin(remaining)]
        n = min(args.solo_per_extra, len(pool))
        ft_n = round(n * (pool.doc_type == "full-text").mean())
        pick = pd.concat([
            stratified_sample(pool[pool.doc_type == "full-text"], ft_n, rng, ["function"]),
            stratified_sample(pool[pool.doc_type == "abstract-only"], n - ft_n, rng, ["function"]),
        ]).drop_duplicates("paper_id")
        sample.loc[sample.paper_id.isin(pick.paper_id), "annotator"] = who
        remaining -= set(pick.paper_id)

    sample = sample.sort_values(["assignment", "order", "doc_type", "function"])
    sample.drop(columns=[c for c in ("notes_blob", "search_text") if c in sample.columns],
                errors="ignore").to_csv(OUT_DIR / "sample_selected.csv", index=False)

    write_workbooks(sample)
    write_stratification_report(papers, sample, args)

    kb = (sample.assignment == "kappa_block").sum()
    ft = (sample.doc_type == "full-text").sum()
    per = " | ".join(f"{ANNOTATORS[k]} {(sample.annotator == k).sum()}"
                     for k in ["bea", *EXTRA_REVIEWERS])
    print(f"\nSample: {len(sample)} papers ({ft} full-text -> paired abstract+full-text "
          f"rows) | kappa block {kb} (all {len(ANNOTATORS)} annotators) | {per}")
    print(f"Outputs in {OUT_DIR}")
    return 0


def expand_rows(block: pd.DataFrame) -> pd.DataFrame:
    """One row per (paper, reading stage). Full-text papers -> an abstract-only row
    then a full-text row (adjacent). Abstract-only papers -> a single abstract row."""
    out = []
    for _, r in block.iterrows():
        if r["doc_type"] == "full-text":
            a = r.copy(); a["reading_stage"] = "abstract"; a["stage_suffix"] = "a"
            b = r.copy(); b["reading_stage"] = "full text"; b["stage_suffix"] = "b"
            out += [a, b]
        else:
            r = r.copy(); r["reading_stage"] = "abstract"; r["stage_suffix"] = ""
            out.append(r)
    return pd.DataFrame(out).reset_index(drop=True)


def write_workbooks(sample: pd.DataFrame) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    HEADER = Font(bold=True, color="FFFFFF", size=10)
    HEADER_FILL = PatternFill("solid", start_color="2C5F8A")
    METHOD_FILL = PatternFill("solid", start_color="E7E0EF")
    SLOT_A = PatternFill("solid", start_color="EAF1F8")
    SLOT_B = PatternFill("solid", start_color="F3ECDA")
    WRAP = Alignment(wrap_text=True, vertical="top")

    Q1 = "Q1  does this paper report a fungus coming from a plant?"
    Q2 = "Q2  paper's main purpose"
    Q3 = "Q3  how were the fungi obtained in this paper?"
    Q4 = "Q4  language of the text you are reading"
    Q5 = "Q5  any sequence-only (uncultured) taxa?"
    Q6 = "Q6  strain-level variation noted?"
    Q7 = "Q7  where was the plant SAMPLED? (not the authors' institution)"
    Q8 = "Q8  what geographic information does the paper contain?"
    Q9 = "Q9  biome (WWF category)"
    Q10 = "Q10  how many distinct fungus-host pairs?"
    Q11 = "Q11  was the surface-sterilisation checked?"

    left_cols = [
        ("row_id", 8), ("paper_reviewed", 22), ("date_reviewed", 12),
        ("reading_stage", 12), ("paper_id", 22), ("doi_link", 20),
        ("year", 6), ("journal", 18), ("title", 40), ("abstract", 66),
        (Q1, 15), (Q2, 20), (Q3, 22), (Q4, 13), (Q5, 15), (Q6, 15),
        (Q7, 20), (Q8, 22), (Q9, 20), (Q10, 12), (Q11, 22),
    ]
    slot_cols = ["fungus", "host_plant", "tissue",
                 "fungal_lifestyle", "effect_on_host", "evidence_basis"]
    right_cols = [
        ("extra_pairs_or_NGS_community_summary", 44),
        ("anything_you_were_unsure_about", 30),
    ]
    dv = {
        "paper_reviewed": "paper_reviewed",
        Q1: "yn_unclear", Q2: "primary_aim", Q3: "sampling_approach",
        Q4: "language", Q5: "yn_unclear", Q6: "strain_variation",
        Q8: "sampling_location_status", Q9: "biome",
        Q11: "sterilisation_checked",
        "tissue": "tissue", "fungal_lifestyle": "fungal_lifestyle",
        "effect_on_host": "effect_on_host", "evidence_basis": "evidence_basis",
    }
    for m in METHOD_COLS:
        dv[m] = "yn"
    q_headers = [Q1, Q2, Q3, Q4, Q5, Q6, Q7, Q8, Q9, Q10, Q11]

    def add_lists_sheet(wb):
        ws = wb.create_sheet("Lists")
        ws.sheet_state = "hidden"
        ranges = {}
        for col, (key, vals) in enumerate(VOCAB.items(), start=1):
            letter = get_column_letter(col)
            ws.cell(1, col, key)
            for i, val in enumerate(vals, start=2):
                ws.cell(i, col, val)
            ranges[key] = f"Lists!${letter}$2:${letter}${len(vals) + 1}"
        return ranges

    def build_sheet(ws, rows: pd.DataFrame, list_ranges):
        headers = [c for c, _ in left_cols] + list(METHOD_COLS)
        for s in range(1, N_SLOTS + 1):
            headers += [f"{c}_{s}" for c in slot_cols]
        headers += [c for c, _ in right_cols]
        ws.append(headers)

        widths = {c: w for c, w in left_cols + right_cols}
        slot_w = {"fungus": 20, "host_plant": 20, "tissue": 15,
                  "fungal_lifestyle": 26, "effect_on_host": 26, "evidence_basis": 26}
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci)
            cell.font, cell.fill, cell.alignment = HEADER, HEADER_FILL, WRAP
            base = h.rsplit("_", 1)[0] if h[-1].isdigit() else h
            ws.column_dimensions[get_column_letter(ci)].width = (
                widths.get(h) or slot_w.get(base) or (19 if h in METHOD_COLS else 13))
        ws.row_dimensions[1].height = 60

        n_q_blanks = sum(1 for c, _ in left_cols if c.startswith("Q"))
        for _, r in rows.iterrows():
            q_vals = [""] * n_q_blanks
            q_vals[3] = str(r.get("language_hint", "") or "")  # Q4 pre-fill hint
            ws.append([
                r["row_id"], "", "", r["reading_stage"],
                r["paper_id"], r["doi_link"], r.get("year", ""),
                str(r.get("journal", ""))[:120], str(r.get("title", ""))[:300],
                str(r.get("abstract", "") or "")[:5000],
            ] + q_vals + [""] * len(METHOD_COLS)
              + [""] * (N_SLOTS * len(slot_cols)) + ["", ""])

        n = len(rows) + 1
        n_left = len(left_cols)
        for ci in range(n_left + 1, n_left + len(METHOD_COLS) + 1):
            col = get_column_letter(ci)
            for rr in range(2, n + 1):
                ws[f"{col}{rr}"].fill = METHOD_FILL
        for si in range(N_SLOTS):
            fill = SLOT_A if si % 2 == 0 else SLOT_B
            c0 = n_left + len(METHOD_COLS) + si * len(slot_cols) + 1
            for ci in range(c0, c0 + len(slot_cols)):
                col = get_column_letter(ci)
                for rr in range(2, n + 1):
                    ws[f"{col}{rr}"].fill = fill

        for ci, h in enumerate(headers, 1):
            base = h.rsplit("_", 1)[0] if h and h[-1].isdigit() else h
            key = dv.get(h) or dv.get(base)
            if not key:
                continue
            # range-referenced (values may contain commas) + "warning" style so an
            # annotator can also type a value that isn't in the list.
            v = DataValidation(type="list", formula1=list_ranges[key], allow_blank=True,
                               showDropDown=False, showErrorMessage=True,
                               errorStyle="warning",
                               error="Not in the list - OK to keep if you mean it.",
                               errorTitle="Custom value")
            ws.add_data_validation(v)
            col = get_column_letter(ci)
            v.add(f"{col}2:{col}{n}")

        abs_col = get_column_letter([c for c, _ in left_cols].index("abstract") + 1)
        for rr, row in enumerate(ws.iter_rows(min_row=2, max_row=n), start=2):
            for cell in row:
                cell.alignment = WRAP
            ws[f"{abs_col}{rr}"].font = Font(size=10)
            ws.row_dimensions[rr].height = 150
        ws.freeze_panes = "E2"  # row_id | paper_reviewed | date_reviewed | reading_stage
        ws.title = "Review"

    def instructions_sheet(ws, who: str, is_kappa: bool, calib_cut: str = ""):
        L = [
            ("Fungal endophyte literature - extraction check", True),
            ("", False), (f"File for: {who}", True), ("", False),
        ]
        if is_kappa and calib_cut:
            L += [
                ("*** DO THE CALIBRATION ROUND FIRST ***", True),
                (f"Complete rows down to and including {calib_cut}, then STOP and tell", False),
                ("Bea. We will compare everyone's answers, sort out any disagreements,", False),
                ("update the rubric if needed, and only then continue with the rest.", False),
                ("", False),
            ]
        L += [
            ("WHAT TO DO", True),
            ("1. Work top to bottom, one row at a time, filling left to right.", False),
            ("2. Set 'paper_reviewed' when you finish a row. It decides how much else you fill:", False),
            ("   complete                     -> fill the whole row from the paper's own data.", False),
            ("   could not access the paper    -> a full-text row you couldn't get. Stop there.", False),
            ("   review or secondary compilation -> Q1 = no, Q3 = review, answer Q4 (language),", False),
            ("        leave Q2, Q5-Q11, the method ticks and the interaction blocks BLANK.", False),
            ("   not a fungus-in-plant study  -> Q1 = no, leave EVERYTHING else blank.", False),
            ("   (A blank 'paper_reviewed' = you haven't done that row yet.)", False),
            ("3. 'reading_stage' tells you how much to read for that row:", False),
            ("     abstract   = read ONLY the abstract text in the 'abstract' column of", False),
            ("                  that row. Do NOT open the link. Answer from the abstract", False),
            ("                  alone, even if that means 'not stated' a lot.", False),
            ("     full text  = open doi_link and read the whole paper", False),
            ("   Full-text papers get TWO adjacent rows: an 'abstract' row (do it FIRST,", False),
            ("   from the abstract alone) then a 'full text' row for the same paper. On the", False),
            ("   'full text' row you may copy your abstract answers across and then correct.", False),
            ("4. Answer Q1-Q11. If the right answer isn't in a dropdown, type it into the", False),
            ("   cell (Excel warns but keeps it). Two that trip people up:", False),
            ("     Q4  language - answer for the TEXT YOU ARE READING on this row. An", False),
            ("         English abstract of a French paper = English on the 'abstract' row,", False),
            ("         French on the 'full text' row.", False),
            ("     Q7/Q8 - Q7 is the SAMPLING location only. If the paper doesn't say where", False),
            ("         sampling was done, write 'not stated' - NEVER put the authors'", False),
            ("         institution in Q7. Q8 then records that only affiliation is available.", False),
            ("5. Review / commentary paper (no new fungi obtained)? Set Q3 = 'review /", False),
            ("   synthesis / commentary' and leave the interaction blocks blank.", False),
            ("6. Method columns (the purple block): put 'yes' in every method the paper", False),
            ("   used to show the fungus was inside the plant - as many as apply. The", False),
            ("   rubric defines each one.", False),
            ("7. List each DISTINCT fungus x host-plant pair the paper NAMES, one per shaded", False),
            ("   block. Each block has SIX cells about that pair:", False),
            ("     fungus / host    - copy the name EXACTLY as the authors write it. If you", False),
            ("                        know it's an old name, still use the authors' name -", False),
            ("                        the update to current taxonomy is done in code later.", False),
            ("     tissue           - if several, put them in one cell: 'leaf, root'.", False),
            ("                        'shoots' / 'aerial parts' with no split -> 'leaf, stem'", False),
            ("     fungal_lifestyle - how the fungus lives (endophyte, saprotroph,", False),
            ("                        mycorrhizal, plant pathogen ...) - see rubric", False),
            ("     effect_on_host   - what happened to THIS plant (harmful / beneficial /", False),
            ("                        neutral / no symptoms not tested / not reported)", False),
            ("     evidence_basis   - was that tested here, observed here, or just inferred", False),
            ("   Leave unused blocks blank.", False),
            ("8. NGS / metabarcoding paper (Q3)? Don't enumerate every OTU - fill blocks", False),
            ("   only for named taxa; summarise the rest in the summary column.", False),
            ("9. Put today's date in 'date_reviewed' (3rd column - no scrolling; type once", False),
            ("   and drag down for a batch). Your name is the file name, so no initials.", False),
            ("   A 'review' or 'not a fungus-in-plant study' row is done entirely in the", False),
            ("   first few columns - you never scroll right for those.", False),
            ("", False),
            ("RULES", True),
            ("- Record what the PAPER (at that reading stage) says, not what you", False),
            ("  believe is biologically true.", False),
            ("- fungal_lifestyle vs effect_on_host are DIFFERENT things (a plant pathogen", False),
            ("  can be present with 'no symptoms'). Both, plus biome and Q11, are defined", False),
            ("  in guild_rubric.md - keep it open beside this file.", False),
            ("- You are NOT checking software output - this is a fresh independent read.", False),
        ]
        if is_kappa:
            L += [("", False),
                  ("This is the shared agreement block. Work independently - do not", False),
                  ("discuss the papers with the other reviewers until all files are in.", False)]
        for i, (t, b) in enumerate(L, 1):
            ws.cell(row=i, column=1, value=t).font = Font(bold=b, size=13 if b and i == 1 else 10)
        ws.column_dimensions["A"].width = 100
        ws.title = "Instructions"

    def assign_ids(block: pd.DataFrame, prefix: str, is_kappa: bool) -> pd.DataFrame:
        # kappa block: order by the stratified sequence so the calibration papers are the
        # first rows. Solo blocks: order by doc type / function for a tidy read.
        sort_by = ["order"] if is_kappa else ["doc_type", "function"]
        block = block.sort_values(sort_by).reset_index(drop=True)
        block["row_id"] = [f"{prefix}%03d" % (i + 1) for i in range(len(block))]
        rows = expand_rows(block)
        rows["row_id"] = rows["row_id"] + rows["stage_suffix"]
        return rows

    def make_wb(block, prefix, who, path, is_kappa):
        wb = Workbook()
        rows = assign_ids(block, prefix, is_kappa)
        calib_cut = ""
        if is_kappa:
            cal = rows[rows["phase"] == "calibration"]
            if len(cal):
                calib_cut = str(cal["row_id"].iloc[-1])
        instructions_sheet(wb.active, who, is_kappa, calib_cut)
        list_ranges = add_lists_sheet(wb)
        build_sheet(wb.create_sheet("Review"), rows, list_ranges)
        try:
            wb.save(OUT_DIR / path)
        except PermissionError:
            alt = OUT_DIR / (Path(path).stem + "_NEW.xlsx")
            wb.save(alt)
            print(f"  WARNING: {path} is open/locked - wrote {alt.name} instead. "
                  f"Close the old file and rename.")

    kappa = sample[sample.assignment == "kappa_block"].copy()
    n_kb_rows = len(expand_rows(kappa))
    for key, label in ANNOTATORS.items():
        make_wb(kappa, "K",
                f"{label} - shared agreement block ({len(kappa)} papers, {n_kb_rows} rows)",
                f"groundtruth_kappa_{key}.xlsx", is_kappa=True)

    for key in ("bea", *EXTRA_REVIEWERS):
        block = sample[(sample.assignment == "solo") & (sample.annotator == key)].copy()
        make_wb(block, key[0].upper(),
                f"{ANNOTATORS[key]} - main block ({len(block)} papers)",
                f"groundtruth_{key}_main.xlsx", is_kappa=False)

    mock = pd.concat([kappa[kappa.doc_type == "abstract-only"].head(2),
                      kappa[kappa.doc_type == "full-text"].head(2)])
    make_wb(mock, "M", "MOCK - 4 example papers (2 abstract-only, 2 full-text)",
            "groundtruth_MOCK.xlsx", is_kappa=False)


def write_stratification_report(papers, sample, args) -> None:
    ft = int((sample.doc_type == "full-text").sum())
    L = [
        "# Ground-truth sample - stratification report", "",
        f"- extraction frame: `{args.extraction.name}`  |  seed: `{args.seed}`",
        f"- core n = {args.n_core} (full-text fraction {args.full_text_frac}) "
        f"+ rare-function oversample n = {args.n_rare}",
        f"- kappa block n = {args.n_kappa} papers, rated by ALL annotators; it is the "
        f"balanced prefix of a stratified pool, so raising --n-kappa later stays valid",
        f"- within the kappa block, the first **{args.n_calibration}** papers are a "
        f"calibration round: everyone does them, then a reconciliation call refines the "
        f"rubric before continuing. Reported Fleiss' kappa uses the remaining "
        f"{args.n_kappa - args.n_calibration} measurement papers (calibration agreement "
        f"reported separately as a pilot)",
        f"- **{len(sample)} unique papers total**; the {ft} full-text papers each get a "
        f"paired abstract-only row + full-text row (~{len(sample) + ft} annotation rows)",
        "", "## Selected sample by stratum", "",
    ]
    for col in ["doc_type", "function", "continent", "era", "assignment", "annotator", "bucket"]:
        if col not in sample.columns:
            continue
        L.append(f"### {col}")
        fv = papers[col].value_counts() if col in papers.columns else None
        if fv is not None:
            L += ["| level | sampled | frame % |", "|---|---|---|"]
            for lvl, k in sample[col].value_counts().sort_index().items():
                L.append(f"| {lvl} | {k} | {100 * int(fv.get(lvl, 0)) / len(papers):.1f}% |")
        else:
            L += ["| level | sampled |", "|---|---|"]
            for lvl, k in sample[col].value_counts().sort_index().items():
                L.append(f"| {lvl} | {k} |")
        L.append("")
    # workload: abstract-only paper = 1 abstract row (~8 min); full-text paper =
    # an abstract row (~5 min) + a full-text row (~20 min).
    AB, FT_A, FT_B = 8 / 60, 5 / 60, 20 / 60

    def load(df):
        n_ab = int((df.doc_type == "abstract-only").sum())
        n_ft = int((df.doc_type == "full-text").sum())
        return n_ab, n_ft, n_ab + 2 * n_ft, n_ab * AB + n_ft * (FT_A + FT_B)

    kb = sample[sample.assignment == "kappa_block"]
    kb_ab, kb_ft, kb_rows, kb_hrs = load(kb)
    L += ["## Workload per annotator (kappa block + own solo block)", "",
          "| annotator | kappa papers | own solo papers | total papers | rows | est. hours |",
          "|---|---|---|---|---|---|"]
    for key in ANNOTATORS:
        solo = sample[(sample.assignment == "solo") & (sample.annotator == key)]
        s_ab, s_ft, s_rows, s_hrs = load(solo)
        L.append(f"| {ANNOTATORS[key]} | {len(kb)} | {len(solo)} | {len(kb) + len(solo)} "
                 f"| {kb_rows + s_rows} | ~{kb_hrs + s_hrs:.0f} |")
    L += ["",
          f"The kappa block is {len(kb)} papers ({kb_ab} abstract-only + {kb_ft} full-text "
          f"= {kb_rows} rows, ~{kb_hrs:.0f} h) - EVERY annotator rates it.",
          "Nancy & Kitty rate only the kappa block.",
          "Each full-text paper = an abstract-only row (done first, ~5 min) + a full-text",
          "row (~20 min); an abstract-only paper = one row (~8 min).",
          "The paired rows give a human-annotated measure of what abstracts omit vs their",
          "own full text - it strengthens the Task 1 abstract/full-text validation."]
    (OUT_DIR / "stratification_report.md").write_text("\n".join(L), encoding="utf-8")


if __name__ == "__main__":
    sys.exit(main())
