#!/usr/bin/env python3
# BMB 2026-06-24
# Builds the supplementary Excel file for the paper — understudied countries,
# plant families, and plant genera (defined as n < 5 studies).

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

RESULTS = "results"
GBIF_TAXON = "data/Reference_datasets/gbif_backbone/Taxon.tsv"
OUT = f"{RESULTS}/supplementary_tables.xlsx"

# ISO A3 -> continent lookup
CONTINENT = {
    # Africa
    "DZA":"Africa","AGO":"Africa","BEN":"Africa","BWA":"Africa","BFA":"Africa",
    "BDI":"Africa","CMR":"Africa","CPV":"Africa","CAF":"Africa","TCD":"Africa",
    "COM":"Africa","COD":"Africa","COG":"Africa","CIV":"Africa","DJI":"Africa",
    "EGY":"Africa","GNQ":"Africa","ERI":"Africa","ETH":"Africa","GAB":"Africa",
    "GMB":"Africa","GHA":"Africa","GIN":"Africa","GNB":"Africa","KEN":"Africa",
    "LSO":"Africa","LBR":"Africa","LBY":"Africa","MDG":"Africa","MWI":"Africa",
    "MLI":"Africa","MRT":"Africa","MUS":"Africa","MAR":"Africa","MOZ":"Africa",
    "NAM":"Africa","NER":"Africa","NGA":"Africa","RWA":"Africa","STP":"Africa",
    "SEN":"Africa","SLE":"Africa","SOM":"Africa","ZAF":"Africa","SSD":"Africa",
    "SDN":"Africa","SWZ":"Africa","TZA":"Africa","TGO":"Africa","TUN":"Africa",
    "UGA":"Africa","ZMB":"Africa","ZWE":"Africa","SHN":"Africa","REU":"Africa",
    "MYT":"Africa","ESH":"Africa",
    # Asia
    "AFG":"Asia","ARM":"Asia","AZE":"Asia","BHR":"Asia","BGD":"Asia","BTN":"Asia",
    "BRN":"Asia","KHM":"Asia","CHN":"Asia","CYP":"Asia","GEO":"Asia","IND":"Asia",
    "IDN":"Asia","IRN":"Asia","IRQ":"Asia","ISR":"Asia","JPN":"Asia","JOR":"Asia",
    "KAZ":"Asia","KWT":"Asia","KGZ":"Asia","LAO":"Asia","LBN":"Asia","MYS":"Asia",
    "MDV":"Asia","MNG":"Asia","MMR":"Asia","NPL":"Asia","PRK":"Asia","OMN":"Asia",
    "PAK":"Asia","PSE":"Asia","PHL":"Asia","QAT":"Asia","SAU":"Asia","SGP":"Asia",
    "KOR":"Asia","LKA":"Asia","SYR":"Asia","TWN":"Asia","TJK":"Asia","THA":"Asia",
    "TLS":"Asia","TUR":"Asia","TKM":"Asia","ARE":"Asia","UZB":"Asia","VNM":"Asia",
    "YEM":"Asia","HKG":"Asia","MAC":"Asia","XKX":"Asia",
    # Europe
    "ALB":"Europe","AND":"Europe","AUT":"Europe","BLR":"Europe","BEL":"Europe",
    "BIH":"Europe","BGR":"Europe","HRV":"Europe","CZE":"Europe","DNK":"Europe",
    "EST":"Europe","FIN":"Europe","FRA":"Europe","DEU":"Europe","GRC":"Europe",
    "HUN":"Europe","ISL":"Europe","IRL":"Europe","ITA":"Europe","LVA":"Europe",
    "LIE":"Europe","LTU":"Europe","LUX":"Europe","MLT":"Europe","MDA":"Europe",
    "MCO":"Europe","MNE":"Europe","NLD":"Europe","MKD":"Europe","NOR":"Europe",
    "POL":"Europe","PRT":"Europe","ROU":"Europe","RUS":"Europe","SMR":"Europe",
    "SRB":"Europe","SVK":"Europe","SVN":"Europe","ESP":"Europe","SWE":"Europe",
    "CHE":"Europe","UKR":"Europe","GBR":"Europe","VAT":"Europe","GGY":"Europe",
    "IMN":"Europe","JEY":"Europe","FRO":"Europe","GIB":"Europe","ALA":"Europe",
    "SJM":"Europe",
    # North America
    "ATG":"North America","BHS":"North America","BRB":"North America",
    "BLZ":"North America","BMU":"North America","CAN":"North America",
    "CRI":"North America","CUB":"North America","DMA":"North America",
    "DOM":"North America","SLV":"North America","GRD":"North America",
    "GTM":"North America","HTI":"North America","HND":"North America",
    "JAM":"North America","MEX":"North America","NIC":"North America",
    "PAN":"North America","KNA":"North America","LCA":"North America",
    "VCT":"North America","TTO":"North America","USA":"North America",
    "PRI":"North America","VIR":"North America","GUM":"North America",
    "CYM":"North America","TCA":"North America","ABW":"North America",
    "CUW":"North America","GLP":"North America","MTQ":"North America",
    "MSR":"North America","AIA":"North America","VGB":"North America",
    "BLM":"North America","MAF":"North America","SPM":"North America",
    "SXM":"North America","BES":"North America",
    # South America
    "ARG":"South America","BOL":"South America","BRA":"South America",
    "CHL":"South America","COL":"South America","ECU":"South America",
    "GUY":"South America","PRY":"South America","PER":"South America",
    "SUR":"South America","URY":"South America","VEN":"South America",
    "GUF":"South America","FLK":"South America",
    # Oceania
    "AUS":"Oceania","FJI":"Oceania","KIR":"Oceania","MHL":"Oceania",
    "FSM":"Oceania","NRU":"Oceania","NZL":"Oceania","PLW":"Oceania",
    "PNG":"Oceania","WSM":"Oceania","SLB":"Oceania","TON":"Oceania",
    "TUV":"Oceania","VUT":"Oceania","NCL":"Oceania","PYF":"Oceania",
    "ASM":"Oceania","MNP":"Oceania","UMI":"Oceania","HMD":"Oceania",
    "CCK":"Oceania","CXR":"Oceania","IOT":"Oceania","ATF":"Oceania",
    "PCN":"Oceania","NFK":"Oceania","COK":"Oceania","NIU":"Oceania",
    "TKL":"Oceania","WLF":"Oceania",
    # Antarctica
    "ATA":"Antarctica",
}

# ── build family->phylum and genus->family,phylum from GBIF backbone ───────
print("Reading GBIF backbone (this takes a moment)...")
gbif = pd.read_csv(
    GBIF_TAXON, sep="\t", usecols=["canonicalName","taxonRank","taxonomicStatus","kingdom","phylum","family","genus"],
    dtype=str, low_memory=False
)
gbif = gbif[gbif["kingdom"] == "Plantae"].copy()
gbif["taxonRank"] = gbif["taxonRank"].str.lower().str.strip()
gbif["taxonomicStatus"] = gbif["taxonomicStatus"].str.lower().str.strip()
gbif["canonicalName"] = gbif["canonicalName"].str.strip()

def modal(series):
    m = series.mode()
    return m.iloc[0] if len(m) > 0 else ""

# family -> phylum: prefer accepted family-rank rows, fall back to phylum from accepted species
fam_rows = gbif[gbif["taxonRank"] == "family"][["canonicalName","phylum"]].dropna()
fam_to_phylum = fam_rows.groupby("canonicalName")["phylum"].agg(modal).to_dict()

# Also build from accepted species in case family-rank rows are missing
sp_accepted = gbif[(gbif["taxonRank"] == "species") & (gbif["taxonomicStatus"] == "accepted")]
sp_fam = sp_accepted[["family","phylum"]].dropna()
for fam, grp in sp_fam.groupby("family"):
    if fam not in fam_to_phylum:
        fam_to_phylum[fam] = modal(grp["phylum"])

# genus -> family, phylum: prefer accepted genus-rank rows, fall back to accepted species
gen_rank = gbif[gbif["taxonRank"] == "genus"][["canonicalName","family","phylum"]].dropna(subset=["canonicalName"])
gen_to_fam    = gen_rank.groupby("canonicalName")["family"].agg(modal).to_dict()
gen_to_phylum = gen_rank.groupby("canonicalName")["phylum"].agg(modal).to_dict()

# fill gaps from accepted species genus column
sp_gen = sp_accepted[["genus","family","phylum"]].dropna(subset=["genus"])
for gen, grp in sp_gen.groupby("genus"):
    if gen not in gen_to_fam and grp["family"].notna().any():
        gen_to_fam[gen]    = modal(grp["family"].dropna())
        gen_to_phylum[gen] = modal(grp["phylum"].dropna())

print(f"  Family->phylum mappings: {len(fam_to_phylum)}")
print(f"  Genus->family mappings:  {len(gen_to_fam)}")

# ── load pipeline results ──────────────────────────────────────────────────
metrics  = pd.read_csv(f"{RESULTS}/country_analysis/geographic_bias_metrics.csv")
families = pd.read_csv(f"{RESULTS}/understudied_analysis/unstudied_plant_families.csv")
genera   = pd.read_csv(f"{RESULTS}/understudied_analysis/unstudied_plant_genera.csv")

# countries: filter to n < 5, add continent, drop study count
countries = metrics[metrics["study_count"] < 5].copy()
countries["continent"] = countries["iso_a3"].map(CONTINENT).fillna("Other")
countries["gdp_current_usd"] = pd.to_numeric(countries["gdp_current_usd"], errors="coerce")
countries = countries.sort_values(["continent", "country_name"])

# families: add phylum, sort by phylum then family
families["phylum"] = families["family"].map(fam_to_phylum).fillna("")
families = families.sort_values(["phylum", "family"])

# genera: add family and phylum, sort by phylum, family, genus
genera["family"] = genera["genus"].map(gen_to_fam).fillna("")
genera["phylum"] = genera["genus"].map(gen_to_phylum).fillna("")
genera = genera.sort_values(["phylum", "family", "genus"])

# ── style helpers ──────────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", start_color="2C5F8A")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BODY_FONT   = Font(name="Arial", size=10)
ALT_FILL    = PatternFill("solid", start_color="EAF1F8")

def style_header(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

def style_body(ws, start_row, end_row, cols):
    for r in range(start_row, end_row + 1):
        fill = ALT_FILL if r % 2 == 0 else None
        for c in range(1, cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = BODY_FONT
            cell.alignment = Alignment(horizontal="left", vertical="center")
            if fill:
                cell.fill = fill

def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ── workbook ───────────────────────────────────────────────────────────────
wb = Workbook()

# ── Notes sheet ────────────────────────────────────────────────────────────
notes_ws = wb.active
notes_ws.title = "Notes"

note_lines = [
    ("Supplementary Tables — Understudied Taxa and Countries", True, 14),
    ("", False, 10),
    ("Definition: 'Understudied' is defined as fewer than 5 publications in the endophyte literature dataset (n < 5).", False, 10),
    ("", False, 10),
    ("Sheet descriptions:", True, 10),
    ("  S1. Understudied Countries — Countries with fewer than 5 endophyte studies, with continent and GDP (current USD, World Bank).", False, 10),
    ("  S2. Understudied Plant Families — Plant families with no endophyte studies identified in the dataset, with associated phylum.", False, 10),
    ("  S3. Understudied Plant Genera — Plant genera with no endophyte studies identified in the dataset, with associated family and phylum.", False, 10),
    ("", False, 10),
    ("Note: The species-level list (>390,000 species) is available in the accompanying Zenodo data repository.", False, 10),
    ("", False, 10),
    ("GDP data source: World Bank, accessed 2024.", False, 10),
    ("Taxonomic backbone: GBIF Backbone Taxonomy.", False, 10),
]

for i, (text, bold, size) in enumerate(note_lines, 1):
    cell = notes_ws.cell(row=i, column=1, value=text)
    cell.font = Font(name="Arial", bold=bold, size=size)
    cell.alignment = Alignment(wrap_text=True)
notes_ws.column_dimensions["A"].width = 110
notes_ws.row_dimensions[1].height = 22

# ── S1: Countries ──────────────────────────────────────────────────────────
c_ws = wb.create_sheet("S1. Understudied Countries")
headers = ["Country", "ISO A3", "Continent", "GDP (current USD)"]
c_ws.append(headers)
style_header(c_ws, 1, len(headers))
c_ws.row_dimensions[1].height = 28

for _, row in countries.iterrows():
    gdp = row["gdp_current_usd"] if pd.notna(row["gdp_current_usd"]) else ""
    c_ws.append([row["country_name"], row["iso_a3"], row["continent"], gdp])

n = len(countries)
style_body(c_ws, 2, n + 1, len(headers))
for r in range(2, n + 2):
    c_ws.cell(row=r, column=4).number_format = '#,##0'
set_widths(c_ws, [26, 10, 18, 22])
c_ws.freeze_panes = "A2"
c_ws.auto_filter.ref = f"A1:D{n + 1}"

# ── S2: Families ───────────────────────────────────────────────────────────
f_ws = wb.create_sheet("S2. Understudied Plant Families")
headers = ["Plant Family", "Phylum"]
f_ws.append(headers)
style_header(f_ws, 1, len(headers))
f_ws.row_dimensions[1].height = 28

for _, row in families.iterrows():
    f_ws.append([row["family"], row["phylum"]])

n = len(families)
style_body(f_ws, 2, n + 1, len(headers))
set_widths(f_ws, [32, 22])
f_ws.freeze_panes = "A2"
f_ws.auto_filter.ref = f"A1:B{n + 1}"

# ── S3: Genera ─────────────────────────────────────────────────────────────
g_ws = wb.create_sheet("S3. Understudied Plant Genera")
headers = ["Plant Genus", "Family", "Phylum"]
g_ws.append(headers)
style_header(g_ws, 1, len(headers))
g_ws.row_dimensions[1].height = 28

for _, row in genera.iterrows():
    g_ws.append([row["genus"], row["family"], row["phylum"]])

n = len(genera)
style_body(g_ws, 2, n + 1, len(headers))
set_widths(g_ws, [28, 28, 22])
g_ws.freeze_panes = "A2"
g_ws.auto_filter.ref = f"A1:C{n + 1}"

wb.save(OUT)
print(f"\nSaved {OUT}")
print(f"  Countries: {len(countries)}")
print(f"  Families:  {len(families)} ({families['phylum'].ne('').sum()} with phylum matched)")
print(f"  Genera:    {len(genera)} ({genera['family'].ne('').sum()} with family matched, {genera['phylum'].ne('').sum()} with phylum matched)")
